import pandas as pd
import numpy as np
import glob
import os
import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime, timedelta

def file_reader(filename):
    errors=[]
    try:
        file=pd.read_excel(filename)
        print(f"{filename} sucessfully loaded with {len(file)}-rows")
    except FileNotFoundError:
        errors.append(" not found")
        print(f"{filename} not found ")
    except Exception as e:
        errors.append(str(e))
        print(f"other errors in founding file {filename} ")
    if len(errors)>0:
        print(f"{errors} Errors occur in loading file{filename}")
    return file,errors
# def file_cleaner(df):
#     print(df.isna())
#     df=df.dropna(how='all')
#     # if "Sales" not in df.columns:
#     #    if "Amount" in df.columns:
#     #     df = df.rename(columns={"Amount": "Sales"})
#     #    elif "sale" in df.columns:
#     #     df = df.rename(columns={"sale": "Sales"})
#     #    elif "money" in df.columns:
#     #     df = df.rename(columns={"money": "Sales"})
#     # if "Quantity" not in df.columns:
#     #    if "Qty" in df.columns:
#     #     df = df.rename(columns={"Qty": "Quantity"})
#     #    elif "Units" in df.columns:
#     #     df = df.rename(columns={"Units": "Quantity"})
       
#     # if "Product" not in df.columns:
#     #    if "Prod" in df.columns:
#     #     df = df.rename(columns={"Prod": "Product"})
#     #    elif "item" in df.columns:
#     #     df = df.rename(columns={"item": "Product"})
#     rename_map = {
#         # Sales variations
#         'Amount': 'Sales', 'Sale': 'Sales', 'Total': 'Sales', 'money': 'Sales',
#         # Quantity variations
#         'Qty': 'Quantity', 'Units': 'Quantity', 'Unit': 'Quantity',
#         # Product variations
#         'Prod': 'Product', 'Item': 'Product', 'item': 'Product', 'ProductName': 'Product'
#     }
    
#     # Apply rename map
#     df = df.rename(columns=rename_map)
#     df['Sales'] = df['Sales'].fillna(df['Sales'].mean())
#     df['Quantity']=df['Quantity'].fillna(1)
#     df['Date']=pd.to_datetime(df['Date'],errors='coerce')
#     df['Date'] = df['Date'].fillna(pd.Timestamp('2024-01-01'))
#     df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
#     df['Source'] = df['Date'].dt.month_name()
#     df['Product']=df['Product'].fillna("Unknown")
#     return df
def file_cleaner(df):
    if df is None:
        return None
    
    df = df.dropna(how='all')
    
    # Simple column mapping
    if 'Item' in df.columns:
        df = df.rename(columns={'Item': 'Product'})
    if 'Qty' in df.columns:
        df = df.rename(columns={'Qty': 'Quantity'})
    if 'Units' in df.columns:
        df = df.rename(columns={'Units': 'Quantity'})
    if 'Amount' in df.columns:
        df = df.rename(columns={'Amount': 'Sales'})
    
    # Create missing columns
    if 'Sales' not in df.columns:
        df['Sales'] = 0
    if 'Quantity' not in df.columns:
        df['Quantity'] = 1
    if 'Product' not in df.columns:
        df['Product'] = 'Unknown'
    if 'Date' not in df.columns:
        df['Date'] = pd.Timestamp('2024-01-01')
    
    # Fill missing values
    df['Sales'] = df['Sales'].fillna(df['Sales'].mean())
    df['Quantity'] = df['Quantity'].fillna(1)
    df['Product'] = df['Product'].fillna('Unknown')
    
    # Dates
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Date'] = df['Date'].fillna(pd.Timestamp('2024-01-01'))
    
    # Source
    df['Source'] = df['Date'].dt.month_name()
    
    return df
def monthly_summary(df):
    """Create simple monthly summary"""
    if df is None or len(df) == 0:
        return pd.DataFrame()
    
    # Extract month from Date
    df['Month'] = df['Date'].dt.strftime('%Y-%m')
    
    # Simple summary
    summary = df.groupby('Month').agg({
        'Sales': ['sum', 'count', 'mean']
    }).reset_index()
    
    # Flatten column names
    summary.columns = ['Month', 'Total_Sales', 'Order_Count', 'Avg_Sale']
    
    # Format numbers
    summary['Total_Sales'] = summary['Total_Sales'].round(2)
    summary['Avg_Sale'] = summary['Avg_Sale'].round(2)
    
    return summary

excel_files=glob.glob('client_files/*.xlsx')
print(f"Files for Automation:\n{excel_files} ")
All_data=[]
errors_log=[]
for file in excel_files:
    df,err=file_reader(file)
    if df is not None:
       cleaned=file_cleaner(df)
       All_data.append(cleaned)
    if err:
       errors_log.append(err)
if All_data:
   master_df=pd.concat(All_data,ignore_index=True)
error_df = pd.DataFrame(errors_log, columns=["Errors"])

summary=monthly_summary(master_df)
if All_data:
    master_df = pd.concat(All_data, ignore_index=True)
    
    # Show what columns we have
    print(f"Available columns: {list(master_df.columns)}")
    
    # Define what we want to keep
    keep_cols = []
    
    if 'Date' in master_df.columns:
        keep_cols.append('Date')
    if 'Product' in master_df.columns:
        keep_cols.append('Product')
    if 'Sales' in master_df.columns:
        keep_cols.append('Sales')
    if 'Quantity' in master_df.columns:
        keep_cols.append('Quantity')
    if 'Source' in master_df.columns:
        keep_cols.append('Source')
    
    # Filter to only these columns
    master_df = master_df[keep_cols]
    
    # Rename Source to Month if it exists
    if 'Source' in master_df.columns:
        master_df = master_df.rename(columns={'Source': 'Month'})
    
    # Reorder columns
    final_order = ['Date', 'Month', 'Product', 'Sales', 'Quantity']
    existing_cols = [col for col in final_order if col in master_df.columns]
    master_df = master_df[existing_cols]
    
    print(f"✅ Final columns: {list(master_df.columns)}")
    print(f"✅ Total rows: {len(master_df)}")
    # Filter to clean colum
with pd.ExcelWriter('Sales_report_Monthly.xlsx',engine='openpyxl') as writer:
   master_df.to_excel(writer,sheet_name='All_data',index=False)
   summary.to_excel(writer,sheet_name='Monthly_Summary',index=False)
   error_df.to_excel(writer, sheet_name='Error_Log', index=False)
   


#  Part 2 adding style

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt

# After saving the Excel file
output_file = 'Sales_report_Monthly.xlsx'

# Create line chart using matplotlib
if len(summary) > 0:
    plt.figure(figsize=(8, 5))
    plt.plot(summary['Month'], summary['Total_Sales'], marker='o', linewidth=2, color='#1f77b4')
    plt.title('Monthly Sales Trend', fontsize=14, fontweight='bold')
    plt.xlabel('Month', fontsize=12)
    plt.ylabel('Total Sales ($)', fontsize=12)
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('temp_chart.png', dpi=100)
    plt.close()
    
    # Add chart to Excel
    wb = load_workbook(output_file)
    
    # Get Monthly Summary sheet
    if 'Monthly_Summary' in wb.sheetnames:
        ws = wb['Monthly_Summary']
        
        # Add chart image
        img = Image('temp_chart.png')
        img.width = 500
        img.height = 350
        # Place chart next to data (assuming data ends at column D)
        ws.add_image(img, 'F2')
        
        # Add formatting
        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers (row 1)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Add borders to all cells with data
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Auto-fit columns
        for col in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        
        # Format numbers in Total_Sales and Avg_Sale columns
        for row in range(2, ws.max_row + 1):
            # Column B is Total_Sales, Column D is Avg_Sale
            for col in [2, 4]:
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        wb.save(output_file)
        print(" Chart and formatting added to Monthly Summary sheet")
        
    
# Also format All_data sheet
wb = load_workbook(output_file)

if 'All_data' in wb.sheetnames:
    ws = wb['All_data']
    
    # Format headers
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Add borders to data rows
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Auto-fit columns
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    wb.save(output_file)
    print("Formatting added to All_data sheet")

print("\n Complete! Report saved with:")
print("   - Blue headers")
print("   - Borders")
print("   - Line chart in Monthly Summary")
print("   - Auto-fitted columns")
    
